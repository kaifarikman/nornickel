"""Deterministic XLSX tails parser → DiagnosticsReport dict.

ИСТОЧНИК ПРАВДЫ по парсингу хвостов — этот файл. `docs/scripts/gen_diagnostics.py` —
frozen-генератор фикстур: при правках здесь перенести изменения туда и
перегенерировать fixtures (`python3 docs/scripts/gen_diagnostics.py` +
`validate_fixtures.py`).

Ядро эндпоинта POST /diagnose. Формат выхода — DiagnosticsReport из docs/CONTRACTS.md.

Парсинг ЯКОРНЫЙ: структура файлов плавает между фабриками (КГМК/НОФ — одна секция
породных хвостов, ТОФ — породные + пирротиновые), поэтому ищем маркерные строки,
а не фиксированные координаты.

Доменные правила диагностики (группы классов крупности, извлекаемость, правила
диагнозов) приходят параметром `config` из `packs/flotation-v1.yaml`
(секция `diagnosis_config`, грузит service.py); константы ниже — fallback-дефолт
при отсутствии секции в pack.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

# Канонизация минералогических форм: подстрока в ячейке → mineral_form из контракта
MINERAL_FORMS = [
    ("раскрыт", "open_pnt_cp"),
    ("закрыт", "closed_pnt_cp"),
    ("примесь в пирротине", "pyrrhotite_impurity"),
    ("силикат", "silicate_valleriite"),
    ("пирит", "pyrite_other_sulfides"),
    ("миллерит", "millerite"),
]

# Fallback-дефолт diagnosis_config — зеркало packs/flotation-v1.yaml
SIZE_GROUPS = {
    "coarse": ["+125", "-125 +71", "+71", "-71 +45"],
    "medium": ["-45 +20", "-20 +10"],
    "fine":   ["-10"],
}
RECOVERABILITY = {
    "element_28": {"open_pnt_cp", "closed_pnt_cp", "millerite"},
    "element_29": {"open_pnt_cp", "closed_pnt_cp"},
}
DIAGNOSIS_RULES = [
    # порядок важен: первое сработавшее правило
    {"diagnosis": "not_recoverable",     "recoverable": False},
    {"diagnosis": "liberation_deficit",  "mineral_form": "closed_pnt_cp", "size_groups": {"coarse"}},
    {"diagnosis": "slimes_overgrinding", "mineral_form": "open_pnt_cp",   "size_groups": {"fine"}},
    {"diagnosis": "flotation_kinetics",  "mineral_form": "open_pnt_cp",   "size_groups": {"coarse", "medium"}},
]
CHECKSUM_TOLERANCE_PCT = 1.0


@dataclass(frozen=True)
class DiagnosisConfig:
    """Доменные правила диагностики из pack (см. diagnosis_config в yaml)."""

    size_groups: dict[str, list[str]]
    recoverability: dict[str, set[str]]
    rules: list[dict[str, Any]]


DEFAULT_DIAGNOSIS_CONFIG = DiagnosisConfig(
    size_groups=SIZE_GROUPS,
    recoverability=RECOVERABILITY,
    rules=DIAGNOSIS_RULES,
)


def norm(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def as_float(v):
    """Число из ячейки; '#REF!' и мусор → None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if s.startswith("#") or not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def size_group(size_class, config: DiagnosisConfig = DEFAULT_DIAGNOSIS_CONFIG):
    for group, classes in config.size_groups.items():
        if size_class in classes:
            return group
    return "medium"


def classify(element, mineral_form, size_class, config: DiagnosisConfig = DEFAULT_DIAGNOSIS_CONFIG):
    recoverable = mineral_form in config.recoverability[element]
    for rule in config.rules:
        if "recoverable" in rule and rule["recoverable"] != recoverable:
            continue
        if "recoverable" in rule:
            return recoverable, rule["diagnosis"]
        if rule.get("mineral_form") != mineral_form:
            continue
        if size_group(size_class, config) not in rule.get("size_groups", set()):
            continue
        return recoverable, rule["diagnosis"]
    # закрытый в средних/тонких классах — тоже дефицит раскрытия по своей природе
    if mineral_form == "closed_pnt_cp":
        return recoverable, "liberation_deficit"
    return recoverable, "flotation_kinetics"


def parse_tails(
    path,
    factory_id,
    config: DiagnosisConfig = DEFAULT_DIAGNOSIS_CONFIG,
) -> dict[str, Any]:
    """xlsx хвостов → dict DiagnosticsReport (см. CONTRACTS.md)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Итог"] if "Итог" in wb.sheetnames else wb.worksheets[0]
    sheet = ws.title
    data_quality = []

    rows = []  # (row_idx, [(col_idx, raw_value), ...]) только непустые ячейки
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        cells = [(c_idx, v) for c_idx, v in enumerate(row, 1) if v is not None]
        rows.append((r_idx, cells))
        for c_idx, v in cells:
            if isinstance(v, str) and v.strip().startswith("#"):
                data_quality.append({
                    "issue": "ref_error",
                    "location": f"{sheet}!{get_column_letter(c_idx)}{r_idx}",
                    "handling": "treated_as_zero",
                })

    def first_text(cells):
        return norm(cells[0][1]) if cells else ""

    def find_rows(predicate):
        return [(r, cells) for r, cells in rows if cells and predicate(first_text(cells).lower())]

    # --- Секции хвостов: породные всегда; пирротиновые — только ТОФ;
    # «Хвосты отвальные» (именно в таком порядке слов) — сводная секция ТОФ,
    # дублирующая rock+pyrrhotite: её блоки пропускаем, иначе двойной счёт ---
    section_anchors = []
    for r, cells in rows:
        t = first_text(cells).lower()
        if t.startswith("хвосты породные") and len(cells) >= 2:
            section_anchors.append((r, "rock"))
        elif t.startswith("хвосты пирротиновые") and len(cells) >= 2:
            section_anchors.append((r, "pyrrhotite"))
        elif t.startswith("хвосты отвальные") or t.startswith("отвальные хвосты общие"):
            section_anchors.append((r, "combined"))
    # У секционных якорей интересуют только те, после которых идёт таблица классов
    class_table_rows = [r for r, cells in rows if first_text(cells).lower().startswith("класс крупности")]
    sections = []
    for table_r in class_table_rows:
        owner = None
        for anchor_r, name in section_anchors:
            if anchor_r < table_r:
                owner = name
        sections.append((table_r, owner or "rock"))
    section_names = sorted({name for _, name in sections if name != "combined"},
                           key=lambda s: 0 if s == "rock" else 1)

    # --- Totals: строка «Отвальные хвосты». Берём ПОСЛЕДНЮЮ до первой таблицы
    # классов: в ТОФ первая строка — округлённое содержание (0.3%), вторая — фактическое ---
    totals = {"tails_smt": None, "element_28": {}, "element_29": {}}
    first_table_r = class_table_rows[0] if class_table_rows else 10**9
    for r, cells in rows:
        if r >= first_table_r:
            break
        if first_text(cells).lower().startswith("отвальные хвосты"):
            vals = [as_float(v) for _, v in cells[1:]]
            vals = [v for v in vals if v is not None]
            if len(vals) >= 5:
                totals = {
                    "tails_smt": vals[0],
                    "element_28": {"pct": round(vals[1], 4), "tons": round(vals[2], 1)},
                    "element_29": {"pct": round(vals[3], 4), "tons": round(vals[4], 1)},
                }

    # --- Блоки минералогии: заголовок «<класс> мкм | Доля потерь Элемент 28 ...».
    # В КГМК встречается заголовок без «мкм» (например «+71») — принимаем и его,
    # но только если в строке есть «Доля потерь» (отличие от строк таблицы классов) ---
    loss_cells = []
    header_re = re.compile(r"^([+\-][\d\s+\-]*?)\s*мкм$")
    bare_header_re = re.compile(r"^([+\-][\d\s+\-]*?)$")

    def block_header(cells):
        """Имя класса крупности, если строка — заголовок блока минералогии, иначе None."""
        t = first_text(cells)
        m = header_re.match(t)
        if m:
            return norm(m.group(1))
        m = bare_header_re.match(t)
        if m and any("доля потерь" in norm(v).lower() for _, v in cells[1:]):
            return norm(m.group(1))
        return None

    for r, cells in rows:
        size_class = block_header(cells)
        if size_class is None:
            continue
        # к какой секции относится блок; combined-секцию пропускаем целиком
        section = "rock"
        for table_r, name in sections:
            if table_r < r:
                section = name
        if section == "combined":
            continue
        # колонки тоннажа: в строке-заголовке ищем «Элемент 28, т» и «Элемент 29, т»
        ton_cols = {}
        for c_idx, v in cells[1:]:
            text = norm(v).lower()
            if "элемент 28" in text and ", т" in text:
                ton_cols["element_28"] = c_idx
            elif "элемент 29" in text and ", т" in text:
                ton_cols["element_29"] = c_idx
            elif text == "элемент 28, т":
                ton_cols["element_28"] = c_idx
            elif text == "элемент 29, т":
                ton_cols["element_29"] = c_idx
        # если «Доля потерь Элемент 28, %» в колонке N, то тонны в N+1 (проверяем соседа)
        if not ton_cols:
            for c_idx, v in cells[1:]:
                text = norm(v).lower()
                if text.startswith("доля потерь элемент 28"):
                    ton_cols["element_28"] = c_idx + 1
                elif text.startswith("доля потерь элемент 29"):
                    ton_cols["element_29"] = c_idx + 1
        if not ton_cols:
            data_quality.append({
                "issue": "parse_warning",
                "location": f"{sheet}!A{r}",
                "handling": f"skipped block {size_class}: no ton columns found",
            })
            continue

        # строки форм до ПЕРВОГО «Итого (проверка)»; затем только первые
        # «Извлекаемый/Не извлекаемый металл» — дальше идут сводные итоги секции,
        # они не относятся к блоку
        block_totals = {}
        stated_recoverable = {}
        for r2, cells2 in rows:
            if r2 <= r:
                continue
            t2 = first_text(cells2).lower()
            if block_header(cells2) is not None or t2.startswith("класс крупности"):
                break  # начался следующий блок
            row_map = dict(cells2)
            if t2.startswith("итого"):
                if block_totals:
                    break  # второе «Итого» — уже сводка секции
                for el, col in ton_cols.items():
                    block_totals[el] = as_float(row_map.get(col))
                continue
            if t2.startswith("извлекаемый металл"):
                for el, col in ton_cols.items():
                    stated_recoverable[el] = as_float(row_map.get(col))
                continue
            if t2.startswith("не извлекаемый"):
                if stated_recoverable:
                    break  # блок полностью прочитан
                continue
            if t2.startswith("потери") or t2.startswith("свободный слот"):
                continue
            if block_totals:
                continue  # формы после «Итого» блоку не принадлежат
            for key, mineral_form in MINERAL_FORMS:
                if t2.startswith(key):
                    share_col = min(ton_cols.values()) - 1  # колонка доли слева от тоннажа
                    for el, col in ton_cols.items():
                        tons = as_float(row_map.get(col))
                        if tons is None:
                            continue
                        recoverable, diagnosis = classify(el, mineral_form, size_class, config)
                        share = as_float(row_map.get(col - 1))
                        loss_cells.append({
                            "section": section,
                            "size_class": size_class,
                            "mineral_form": mineral_form,
                            "element": el,
                            "tons": round(tons, 2),
                            "share_of_class_pct": round(share, 2) if share is not None else None,
                            "recoverable": recoverable,
                            "diagnosis": diagnosis,
                            "cell_ref": f"{sheet}!{get_column_letter(col)}{r2}",
                        })
                    break

        # самопроверка: сумма форм против «Итого (проверка)»
        for el, stated in block_totals.items():
            if stated is None:
                continue
            got = sum(c["tons"] for c in loss_cells
                      if c["size_class"] == size_class and c["element"] == el and c["section"] == section)
            if stated and abs(got - stated) / stated * 100 > CHECKSUM_TOLERANCE_PCT:
                data_quality.append({
                    "issue": "checksum_mismatch",
                    "location": f"{sheet} класс {size_class}, {el}",
                    "handling": "reported",
                    "delta_pct": round(abs(got - stated) / stated * 100, 2),
                })
        # самопроверка: наша recoverability против строки «Извлекаемый металл»
        for el, stated in stated_recoverable.items():
            if stated is None:
                continue
            got = sum(c["tons"] for c in loss_cells
                      if c["size_class"] == size_class and c["element"] == el
                      and c["section"] == section and c["recoverable"])
            if stated and abs(got - stated) / stated * 100 > CHECKSUM_TOLERANCE_PCT:
                data_quality.append({
                    "issue": "checksum_mismatch",
                    "location": f"{sheet} класс {size_class}, {el} (извлекаемый металл)",
                    "handling": "reported",
                    "delta_pct": round(abs(got - stated) / stated * 100, 2),
                })

    # --- сверка на уровне файла: сумма loss_cells против заявленных totals ---
    for el in ("element_28", "element_29"):
        stated = totals.get(el, {}).get("tons")
        if not stated:
            continue
        got = sum(c["tons"] for c in loss_cells if c["element"] == el)
        delta_pct = abs(got - stated) / stated * 100
        if delta_pct > CHECKSUM_TOLERANCE_PCT:
            data_quality.append({
                "issue": "checksum_mismatch",
                "location": f"файл целиком, {el}: сумма классов {got:,.1f} т vs шапка {stated:,.1f} т",
                "handling": "reported",
                "delta_pct": round(delta_pct, 2),
            })

    # --- diagnosis_summary ---
    summary = {}
    for c in loss_cells:
        key = (c["diagnosis"], c["element"])
        summary[key] = summary.get(key, 0.0) + c["tons"]
    diagnosis_summary = [
        {"diagnosis": d, "element": el, "tons": round(t, 1)}
        for (d, el), t in sorted(summary.items(), key=lambda kv: -kv[1])
    ]

    return {
        "factory_id": factory_id,
        "pack_id": "flotation-v1",
        "source_file": str(path),
        "sections": section_names,
        "totals": totals,
        "loss_cells": loss_cells,
        "diagnosis_summary": diagnosis_summary,
        "data_quality": data_quality,
    }
